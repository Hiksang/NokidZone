from korean_geocoding.geocoding import KoreanGeocoding as Kg
import openpyxl
import time

kg = Kg()
kg.set_naver_api("hdfa8f31bk", "3uD3bopOrcTwYjeC6TnLOQxGDqIkwNPPYlhYo4JI") # 발급받은 키를 넣어야 합니다.
# print(kg.get_coordinates_by_api("서울특별시 종로구 세검정로 430 (평창동,(지하1층)"))
#
#
#
## 엘셀파일 열기
filename = "last.xlsx"
exelFile = openpyxl.load_workbook(filename)

## 시트 설정
sheet = exelFile.worksheets[0]



rowCount = 1

for row in sheet.rows:
    try:
        print(row[1].value)
        g = kg.get_coordinates_by_api(row[1].value)
        lat_cell = sheet.cell(row=rowCount, column=3)
        lng_cell = sheet.cell(row=rowCount, column=4)

        lat_cell.value = g[0]
        lng_cell.value = g[1]
        rowCount = rowCount + 1
    except:
        rowCount = rowCount + 1
        print("error")


exelFile.save("nokids.xlsx")